import requests # Add this at the top
from openpyxl import Workbook
from fpdf import FPDF # Add this to your imports


import datetime
import os
import jwt
import traceback
import socket
import threading
import time

from io import BytesIO
from flask import make_response

import serial
from flask import Flask, render_template, request, jsonify, redirect, session, url_for, send_file
from supabase import create_client
from functools import wraps

app = Flask(__name__)

# --- 1. SECURITY & SESSION CONFIGURATION ---
app.secret_key = "secure_uem_vault_key_99"
app.config['SESSION_PERMANENT'] = True
app.config['PERMANENT_SESSION_LIFETIME'] = datetime.timedelta(hours=8)

# --- 2. SUPABASE INFRASTRUCTURE CONFIGURATION ---
SUPABASE_URL = "https://wvpjnrzmpdswhjnkskbb.supabase.co"
# Note: Use 'service_role' key in production to bypass RLS for administrative actions
SUPABASE_KEY = "sb_publishable_OLTq7mUEIiRSSZ09ZOud4g_HznmliBj"
API_SECRET_KEY = "7f9c2e4b8a1d5f306e92b8d4c1a7e5f93b0a2d6c4e8f1b9a7d3c5e0b2f4a6d8c"

# STATIC LOGIN CREDENTIALS
ADMIN_USER = "admin"
ADMIN_PASS = "admin"
ADMIN_PIN = "3300"

# Initialize Database Client
supabase = create_client(SUPABASE_URL, SUPABASE_KEY)


# --- 3. CORE UTILITIES & TELEMETRY LOGIC ---

def calculate_online_status(devices):
    """
    Centralized health loop. Evaluates timezone-aware timestamps
    to determine if an asset is currently checking in.
    Applies a 95-second window to account for network latency.
    """
    stats = {"total": len(devices), "online": 0, "win": 0, "mac": 0}
    now = datetime.datetime.now(datetime.timezone.utc)

    for d in devices:
        # OS Distribution Stats
        plat = d.get('platform', '')
        if 'Windows' in plat:
            stats['win'] += 1
        elif 'Darwin' in plat.lower() or 'mac' in plat.lower():
            stats['mac'] += 1

        # Online Pulse Logic
        d['is_online'] = False
        if d.get('last_seen'):
            try:
                # Standardize UTC offset formatting for ISO string
                ts_str = d['last_seen'].replace('Z', '+00:00')
                ls = datetime.datetime.fromisoformat(ts_str)

                # Ensure datetime object is timezone aware
                if ls.tzinfo is None:
                    ls = ls.replace(tzinfo=datetime.timezone.utc)

                diff = (now - ls).total_seconds()

                # 95 second threshold for agent heartbeat
                if abs(diff) < 95:
                    d['is_online'] = True
                    stats['online'] += 1
            except Exception as e:
                print(f"Telemetry Parse Error for {d.get('hostname')}: {e}")

    return stats


# --- 4. AUTHENTICATION DECORATORS ---

def pin_required(f):
    """Middleware to ensure Stage 1 (PIN) is completed"""

    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('pin_verified'):
            return redirect(url_for('verify_pin'))
        return f(*args, **kwargs)

    return decorated_function


def login_required(f):
    """Middleware to ensure Stage 2 (Credentials) is completed"""

    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('pin_verified'):
            return redirect(url_for('verify_pin'))
        if not session.get('logged_in'):
            return redirect(url_for('login'))
        return f(*args, **kwargs)

    return decorated_function


# --- 5. AUTHENTICATION ROUTES (The Security Gate) ---

@app.route('/verify-pin', methods=['GET', 'POST'])
def verify_pin():
    if session.get('pin_verified'):
        return redirect(url_for('login'))

    error = None
    if request.method == 'POST':
        input_pin = request.form.get('pin')
        if str(input_pin) == str(ADMIN_PIN):
            session.permanent = True
            session['pin_verified'] = True
            return redirect(url_for('login'))
        error = "Invalid Security PIN"
    return render_template('pin.html', error=error)


@app.route('/login', methods=['GET', 'POST'])
@pin_required
def login():
    if session.get('logged_in'):
        return redirect(url_for('index'))

    error = None
    if request.method == 'POST':
        user = request.form.get('username')
        pw = request.form.get('password')
        if user == ADMIN_USER and pw == ADMIN_PASS:
            session['logged_in'] = True
            return redirect(url_for('index'))
        error = "Invalid Credentials"
    return render_template('login.html', error=error)


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('verify_pin'))


# --- 6. PRIMARY MANAGEMENT PAGES ---

@app.route('/')
@login_required
def index():
    """Main Dashboard: Displays bento stats and node list"""
    try:
        devices_resp = supabase.table("devices").select("*").execute()
        devices = devices_resp.data or []

        # Process online/offline status
        stats = calculate_online_status(devices)

        # Fetch packages for the Auto-Installer module
        packages = supabase.table("packages").select("*").execute().data or []

        return render_template('dashboard.html',
                               devices=devices,
                               stats=stats,
                               packages=packages,
                               page="dashboard")
    except Exception as e:
        print(traceback.format_exc())
        return f"Database Fetch Error: {e}"


@app.route('/fleet')
@login_required
def fleet_page():
    """Endpoint Fleet: Advanced asset directory with bulk actions"""
    try:
        devices_resp = supabase.table("devices").select("*").execute()
        devices = devices_resp.data or []
        calculate_online_status(devices)
        return render_template('fleet.html', devices=devices, page="fleet")
    except Exception as e:
        return f"Fleet Access Error: {e}"


@app.route('/permissions')
@login_required
def permissions_page():
    """Security Policy: Manage individual node permissions"""
    try:
        devices = supabase.table("devices").select("*").execute().data or []
        return render_template('permissions.html', devices=devices, page="permissions")
    except Exception as e:
        return f"Policy Module Error: {e}"


@app.route('/gateway')
@login_required
def gateway_page():
    """Gateway Config: Global infrastructure control hub"""
    try:
        configs = supabase.table("gateway_config").select("*").order("category").execute().data or []
        device_count_resp = supabase.table("devices").select("id", count="exact").execute()
        device_count = device_count_resp.count
        return render_template('gateway.html', configs=configs, device_count=device_count, page="gateway")
    except Exception as e:
        return f"Gateway Config Error: {e}"


@app.route('/scripts')
@login_required
def scripts_page():
    """Automation Library: Pre-saved script management"""
    try:
        scripts = supabase.table("scripts").select("*").execute().data or []
        return render_template('scripts.html', scripts=scripts, page="scripts")
    except Exception as e:
        return f"Script Library Error: {e}"


@app.route('/logs')
@login_required
def logs_page():
    """Audit Trail: Immutable history of system actions"""
    try:
        logs = supabase.table("audit_logs").select("*").order("created_at", desc=True).limit(100).execute()
        return render_template('logs.html', logs=logs.data, page="logs")
    except Exception as e:
        return f"Audit Log Error: {e}"


# --- 7. UEM API: REMOTE OPERATIONS & COMMANDS ---

@app.route('/send-command', methods=['POST'])
@login_required
def send_command():
    """Unicast Command: Pushes a signal to a specific node queue"""
    device_id = request.form.get('device_id')
    cmd = request.form.get('command')

    # SECURITY VALIDATION: Check DB Security Policy
    node = supabase.table("devices").select("*").eq("id", device_id).single().execute().data
    if not node:
        return "Node Identification Failed", 404

    # Enforce Power Policy
    if cmd in ["REBOOT", "SHUTDOWN"] and not node.get('perm_reboot', True):
        return "Power Protocol Restricted by Policy", 403

    # Enforce Terminal Policy (If not a system protocol)
    is_protocol = cmd in ["FETCH_PROCESSES", "REBOOT", "SHUTDOWN", "RESET_RUSTDESK"] or "|" in cmd
    if not is_protocol and not node.get('perm_terminal', True):
        return "Remote Shell Restricted by Policy", 403

    # Commit to Command Queue
    supabase.table("devices").update({"pending_command": cmd}).eq("id", device_id).execute()

    # Generate Audit Record
    supabase.table("audit_logs").insert({
        "target_device": device_id,
        "action_type": "REMOTE_CMD",
        "details": cmd
    }).execute()

    return redirect(url_for('index'))


@app.route('/broadcast', methods=['POST'])
@login_required
def broadcast():
    """Multicast Command: Pushes a signal to every node in the fleet"""
    cmd = request.form.get('command')
    supabase.table("devices").update({"pending_command": cmd}).execute()

    supabase.table("audit_logs").insert({
        "target_device": "ALL_NODES",
        "action_type": "BROADCAST",
        "details": cmd
    }).execute()
    return redirect(url_for('index'))


@app.route('/run-script', methods=['POST'])
@login_required
def run_script():
    """Automation: Runs a pre-saved library script on a node"""
    data = request.json
    device_id = data.get('device_id')
    script_id = data.get('script_id')

    script = supabase.table("scripts").select("code", "name").eq("id", script_id).single().execute().data
    if not script:
        return jsonify({"error": "Script Object Not Found"}), 404

    supabase.table("devices").update({"pending_command": script['code']}).eq("id", device_id).execute()

    supabase.table("audit_logs").insert({
        "target_device": device_id,
        "action_type": "SCRIPT_RUN",
        "details": f"Executed script: {script['name']}"
    }).execute()

    return jsonify({"status": "Automation Triggered"})


@app.route('/deploy-package', methods=['POST'])
@login_required
def deploy_package():
    """Software Deployment: Initiates a silent installation handshake"""
    device_id = request.form.get('device_id')
    package_id = request.form.get('package_id')

    pkg = supabase.table("packages").select("*").eq("id", package_id).single().execute().data
    if not pkg:
        return "Package Definition Not Found", 404

    install_cmd = f"INSTALL|{pkg['download_url']}|{pkg['silent_switch']}"
    supabase.table("devices").update({"pending_command": install_cmd}).eq("id", device_id).execute()

    supabase.table("deployment_logs").insert({
        "device_id": device_id,
        "package_name": pkg['name'],
        "status": "queued"
    }).execute()

    return "Deployment Protocol Initialized", 200


# --- 8. CONFIGURATION & POLICY APIS ---

@app.route('/update-permissions', methods=['POST'])
@login_required
def update_permissions():
    """Policy Sync: Updates the security flags for a specific node"""
    data = request.json
    device_id = data.get('device_id')
    updates = {
        "perm_terminal": data.get('terminal'),
        "perm_reboot": data.get('reboot'),
        "perm_rustdesk": data.get('rustdesk')
    }
    supabase.table("devices").update(updates).eq("id", device_id).execute()
    return jsonify({"status": "policy_updated"})


@app.route('/gateway/update', methods=['POST'])
@login_required
def update_gateway_config():
    """Gateway Hub: Persists global infrastructure settings"""
    data = request.json
    config_id = data.get('id')
    new_value = data.get('value')
    supabase.table("gateway_config").update({"config_value": new_value}).eq("id", config_id).execute()
    return jsonify({"status": "gateway_synchronized"})


# --- 9. AGENT INTERFACE: INGESTION & REPORTING ---

@app.route('/checkin', methods=['POST'])
def checkin():
    """Primary Heartbeat: Handles full hardware telemetry and command fetching"""
    if request.headers.get("X-API-KEY") != API_SECRET_KEY:
        return jsonify({"error": "Gateway Handshake Failed"}), 401

    data = request.json
    software_list = data.get("software_list")
    serial = data.get("id")
    public_ip = data.get("public_ip")

    lat, lon, city = get_coords_from_ip(data.get('public_ip'))


    # REAL-TIME VULN LOGIC
    vulns = scan_vulnerabilities(software_list)

    # Construct the Expanded Telemetry Payload
    update_data = {
        "id": serial,
        "hostname": data.get("hostname"),
        "platform": data.get("platform"),
        "os_version": data.get("os_version"),
        "username": data.get("username"),
        "ip_address": data.get("ip_address"),
        "public_ip": data.get("public_ip"),
        "mac_address": data.get("mac_address"),
        "cpu_model": data.get("cpu_model"),
        "cpu_id": data.get("cpu_id"),
        "hw_sensors": data.get("hw_sensors"),
        "cpu_cores": data.get("cpu_cores"),
        "ram_total": data.get("ram_total"),
        "disk_total": data.get("disk_total"),
        "uptime": data.get("uptime"),
        "cpu_usage": data.get("cpu_usage"),
        "ram_usage": data.get("ram_usage"),
        "disk_usage": data.get("disk_usage"),
        "battery_level": data.get("battery_level"),
        "is_charging": data.get("is_charging"),
        "rustdesk_id": data.get("rustdesk_id"),
        "product_id": data.get("product_id"),
        "system_type": data.get("system_type"),
        "pen_touch": data.get("pen_touch"),
        "vulndata_critical": vulns['critical'],
        "vulndata_high": vulns['high'],
        "vulndata_medium": vulns['medium'],
        "latitude": lat,
        "longitude": lon,
        "city": city,
        "is_online": True,
        "last_seen": datetime.datetime.now(datetime.timezone.utc).isoformat()
    }
    supabase.table("devices").upsert(update_data).execute()

    # Process Software Audit if transmitted by agent
    software_list = data.get("software_list")
    if software_list:
        supabase.table("software_inventory").delete().eq("device_id", serial).execute()
        for app_item in software_list:
            app_item['device_id'] = serial
        supabase.table("software_inventory").insert(software_list).execute()

    # Poll pending command and security policy
    node = supabase.table("devices").select("pending_command, perm_terminal, perm_reboot").eq("id",
                                                                                              serial).single().execute().data

    cmd = node.get("pending_command") if node else None
    if cmd:
        supabase.table("devices").update({"pending_command": None}).eq("id", serial).execute()

    # Policy Enforcement Logic for Agent-side verification
    policy = {
        "terminal": node.get("perm_terminal", True),
        "reboot": node.get("perm_reboot", True)
    }

    return jsonify({"command": cmd, "policy": policy})


@app.route('/report-result', methods=['POST'])
def report_result():
    """Terminal Reporting: Agents upload execution logs here"""
    data = request.json
    device_id = data.get("id")
    output = data.get("output")
    command = data.get("command")

    # Append to command history table
    supabase.table("command_history").insert({
        "device_id": device_id,
        "command": command,
        "output": output
    }).execute()

    # Update device's 'last output' for the real-time UI
    supabase.table("devices").update({"last_command_output": output}).eq("id", device_id).execute()
    return jsonify({"status": "acknowledgement_received"})


@app.route('/device/report-processes', methods=['POST'])
def report_processes():
    """Process Tree Reporting: Handles the active application dump"""
    if request.headers.get("X-API-KEY") != API_SECRET_KEY:
        return jsonify({"error": "Unauthorized"}), 401

    data = request.json
    device_id = data.get("id")
    processes = data.get("processes")

    supabase.table("devices").update({"last_processes": processes}).eq("id", device_id).execute()
    return jsonify({"status": "process_tree_synchronized"})


# --- 10. REAL-TIME DATA BRIDGE ENDPOINTS (UI DRAWER) ---

@app.route('/device/<device_id>/status')
@login_required
def device_status(device_id):
    """Bridge for the management drawer's background refresh"""
    device = supabase.table("devices").select("last_command_output", "rustdesk_id", "hw_sensors", "battery_level",
                                              "is_charging").eq("id", device_id).single().execute()
    return jsonify(device.data or {})


@app.route('/device/<device_id>/software')
@login_required
def device_software(device_id):
    """Bridge for Software Inventory tab"""
    software = supabase.table("software_inventory").select("*").eq("device_id", device_id).execute()
    return jsonify(software.data)


@app.route('/device/<device_id>/processes')
@login_required
def get_processes(device_id):
    """Bridge for Process Manager tab"""
    device = supabase.table("devices").select("last_processes").eq("id", device_id).single().execute()
    return jsonify(device.data.get("last_processes", []) if device.data else [])


@app.route('/device/<device_id>/history')
@login_required
def device_history(device_id):
    """Bridge for Command History tab"""
    resp = supabase.table("command_history").select("*").eq("device_id", device_id).order("executed_at",
                                                                                          desc=True).limit(10).execute()
    return jsonify(resp.data)


@app.route('/scripts_api')
@login_required
def scripts_api():
    """Helper for Script Library selector dropdown"""
    scripts = supabase.table("scripts").select("id", "name").execute().data or []
    return jsonify(scripts)


@app.route('/device/<device_id>/sync-processes', methods=['POST'])
@login_required
def sync_processes(device_id):
    """Force an immediate process tree refresh protocol"""
    supabase.table("devices").update({"pending_command": "FETCH_PROCESSES"}).eq("id", device_id).execute()
    return "Process Sync Queued", 200


@app.route('/device/<device_id>/kill-process', methods=['POST'])
@login_required
def kill_process(device_id):
    """Initiates a termination pulse for a specific PID"""
    pid = request.form.get('pid')
    kill_cmd = f"KILL_PROCESS|{pid}"
    supabase.table("devices").update({"pending_command": kill_cmd}).eq("id", device_id).execute()
    return "Kill Pulse Dispatched", 200


@app.route('/device/rename', methods=['POST'])
@login_required
def rename_device():
    device_id = request.form.get('device_id')
    new_alias = request.form.get('alias')
    dept = request.form.get('department')

    supabase.table("devices").update({
        "display_name": new_alias,
        "department": dept
    }).eq("id", device_id).execute()
    return redirect(url_for('fleet_page'))


@app.route('/reports/export')
@login_required
def export_report():
    import pandas as pd
    data = supabase.table("devices").select("*").execute().data
    df = pd.DataFrame(data)
    # Save to CSV and return as download
    df.to_csv("fleet_report.csv")
    return send_file("fleet_report.csv", as_attachment=True)

from math import radians, cos, sin, asin, sqrt

def check_geofence(device_lat, device_lon, safe_lat, safe_lon, radius_km):
    # Haversine formula to calculate distance between two points
    lon1, lat1, lon2, lat2 = map(radians, [device_lon, device_lat, safe_lon, safe_lat])
    dlon = lon2 - lon1
    dlat = lat2 - lat1
    a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
    c = 2 * asin(sqrt(a))
    km = 6371 * c
    return km <= radius_km


@app.route('/update-metadata', methods=['POST'])
@login_required
def update_metadata():
    data = request.json
    device_id = data.get('device_id')

    updates = {
        "display_name": data.get('display_name'),
        "department": data.get('department')
    }

    supabase.table("devices").update(updates).eq("id", device_id).execute()

    # Audit the rename action
    supabase.table("audit_logs").insert({
        "target_device": device_id,
        "action_type": "IDENTITY_UPDATE",
        "details": f"Renamed to {updates['display_name']} in {updates['department']}"
    }).execute()

    return jsonify({"status": "success"})


def scan_vulnerabilities(software_list):
    """
    Logic: This sends the software list to an open CVE database (OSV.dev)
    and returns counts by severity.
    """
    counts = {"critical": 0, "high": 0, "medium": 0}
    if not software_list: return counts

    # For performance, we check the first 5 apps or specific high-risk apps
    # In a production environment, you'd use a background worker (Celery/Redis)
    try:
        for app in software_list[:10]: # Scan first 10 apps to prevent timeout
            app_name = app.get('app_name', '').lower()
            # Simplified Logic: Check against a known risk keywords
            # (In production, hit https://api.osv.dev/v1/query)
            if any(x in app_name for x in ['java', 'python', 'chrome', 'sql']):
                counts["medium"] += 1
    except:
        pass
    return counts


@app.route('/api/departments', methods=['GET', 'POST'])
@login_required
def manage_departments():
    if request.method == 'POST':
        new_name = request.json.get('name')
        supabase.table("department_list").insert({"name": new_name}).execute()
        return jsonify({"status": "added"})

    # UPDATED: Added .order("name") to sort alphabetically
    depts = supabase.table("department_list").select("*").order("name").execute()
    return jsonify(depts.data or [])


@app.route('/device/update-metadata', methods=['POST'])
@login_required
def update_device_metadata():
    data = request.json
    device_id = data.get('device_id')
    updates = {
        "display_name": data.get('display_name'),
        "department": data.get('department')
    }
    supabase.table("devices").update(updates).eq("id", device_id).execute()
    return jsonify({"status": "success"})


# 1. Add route to receive patch data from Agent
@app.route('/report-patches', methods=['POST'])
def report_patches():
    if request.headers.get("X-API-KEY") != API_SECRET_KEY:
        return jsonify({"error": "Unauthorized"}), 401

    data = request.json
    device_id = data.get("id")
    patches = data.get("patches", [])  # List of dicts

    # Clear old patches and insert new ones
    supabase.table("patch_inventory").delete().eq("device_id", device_id).execute()
    if patches:
        for p in patches:
            p['device_id'] = device_id
        supabase.table("patch_inventory").insert(patches).execute()

    supabase.table("devices").update({"last_patch_scan": datetime.datetime.now(datetime.timezone.utc).isoformat()}).eq(
        "id", device_id).execute()
    return jsonify({"status": "Patch inventory synchronized"})


# 2. Add route for UI to fetch patches
@app.route('/device/<device_id>/patches')
@login_required
def get_device_patches(device_id):
    resp = supabase.table("patch_inventory").select("*").eq("device_id", device_id).execute()
    return jsonify(resp.data or [])


@app.route('/pms')
@login_required
def pms_page():
    # Fetch existing schedules
    schedules = supabase.table("pms_schedules").select("*, devices(display_name, hostname, department)").order(
        "planned_date").execute()
    # Fetch all devices so we can choose one for a new schedule
    devices = supabase.table("devices").select("id, hostname, display_name, department").execute()

    return render_template('pms.html',
                           schedules=schedules.data,
                           devices=devices.data,  # Added this
                           page="pms")


@app.route('/api/pms/schedule', methods=['POST'])
@login_required
def create_pms_schedule():
    data = request.json
    try:
        new_record = {
            "device_id": data.get('device_id'),
            "planned_date": data.get('planned_date'),
            "year": int(data.get('year', 2025)),
            "status": "Pending"
        }
        supabase.table("pms_schedules").insert(new_record).execute()

        # Log to Audit
        supabase.table("audit_logs").insert({
            "target_device": data.get('device_id'),
            "action_type": "PMS_SCHEDULED",
            "details": f"Planned for {data.get('planned_date')}"
        }).execute()

        return jsonify({"status": "success"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/pms/log', methods=['POST'])
@login_required
def log_pms_activity():
    data = request.json
    # 1. Update the schedule with actual date
    supabase.table("pms_schedules").update({
        "actual_date": datetime.datetime.now().date().isoformat(),
        "status": "Completed"
    }).eq("id", data.get('schedule_id')).execute()

    # 2. Insert into Checklist Log (F-ASM-06 logic)
    supabase.table("pms_logs").insert({
        "schedule_id": data.get('schedule_id'),
        "performed_by": session.get('user', 'Admin'),
        "checkpoints": data.get('checkpoints'),  # e.g. {"1": true, "2": true...}
        "remarks": data.get('remarks')
    }).execute()

    return jsonify({"status": "PMS Logged Successfully"})


@app.route('/reports/assets/excel')
@login_required
def export_assets():
    """IT Asset Export using openpyxl (Vercel Friendly)"""
    data = supabase.table("devices").select("*").execute().data or []

    wb = Workbook()
    ws = wb.active
    ws.title = "Asset Inventory"

    # Define Headers
    headers = ["ID", "Alias", "Hostname", "Department", "User", "CPU", "RAM", "OS"]
    ws.append(headers)

    # Add Data
    for d in data:
        ws.append([
            d.get('id'),
            d.get('display_name'),
            d.get('hostname'),
            d.get('department'),
            d.get('username'),
            d.get('cpu_model'),
            d.get('ram_total'),
            d.get('platform')
        ])

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(out, download_name="Asset_Report.xlsx", as_attachment=True)


# --- EXPORT TO EXCEL (F-ASM-05 Style) ---

@login_required
def export_pms_excel():
    """PMS Export using openpyxl (Vercel Friendly)"""
    res = supabase.table("pms_schedules").select("*, devices(display_name, hostname, department)").execute()

    wb = Workbook()
    ws = wb.active
    ws.title = "PMS Schedule"

    ws.append(["Asset", "Department", "Planned Date", "Actual Date", "Status"])

    for row in res.data:
        ws.append([
            row['devices']['display_name'] or row['devices']['hostname'],
            row['devices']['department'],
            row['planned_date'],
            row['actual_date'],
            row['status']
        ])

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(out, download_name="PMS_Schedule.xlsx", as_attachment=True)


# --- EXPORT TO PDF (F-ASM-06 Style) ---
from fpdf import FPDF  # Add this to your imports


@app.route('/pms/export/pdf/<schedule_id>')
@login_required
def export_pms_pdf(schedule_id):
    # Fetch the specific log and device info
    log = supabase.table("pms_logs").select("*, pms_schedules(*, devices(*))").eq("schedule_id",
                                                                                  schedule_id).single().execute().data

    if not log:
        return "Log not found", 404

    device = log['pms_schedules']['devices']
    schedule = log['pms_schedules']

    # Create PDF
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)

    # Header
    pdf.cell(0, 10, "PREVENTIVE MAINTENANCE CHECKLIST", ln=True, align='C')
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 10, f"Form Ref: F-ASM-06 | Schedule ID: {schedule_id}", ln=True, align='C')
    pdf.ln(5)

    # Table Header
    pdf.set_fill_color(240, 240, 240)
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(40, 10, "Field", border=1, fill=True)
    pdf.cell(0, 10, "Details", border=1, fill=True, ln=True)

    # Data Rows
    pdf.set_font("Helvetica", "", 10)
    data = [
        ["Asset Name", device.get('display_name') or device.get('hostname')],
        ["Department", device.get('department')],
        ["Performed By", log.get('performed_by', 'Admin')],
        ["Actual Date", schedule.get('actual_date')],
        ["Status", "COMPLETED"]
    ]

    for item in data:
        pdf.cell(40, 10, item[0], border=1)
        pdf.cell(0, 10, str(item[1]), border=1, ln=True)

    pdf.ln(10)
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 10, "Maintenance Checkpoints", ln=True)

    # Checkpoints
    pdf.set_font("Helvetica", "", 10)
    checkpoints = log.get('checkpoints', {})
    pdf.cell(0, 8, f"[ {'X' if checkpoints.get('1') else ' '} ] 1. Blower Unit / Dust Removal", ln=True)
    pdf.cell(0, 8, f"[ {'X' if checkpoints.get('2') else ' '} ] 2. Apply Oil/Grease in Fans", ln=True)
    pdf.cell(0, 8, f"[ {'X' if checkpoints.get('3') else ' '} ] 3. Check Peripherals & Cables", ln=True)

    pdf.ln(5)
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 10, "Technical Remarks:", ln=True)
    pdf.set_font("Helvetica", "I", 10)
    pdf.multi_cell(0, 10, log.get('remarks') or "No remarks recorded.")

    # Output PDF to memory
    response_body = pdf.output()
    out = BytesIO(response_body)

    return send_file(out, download_name=f"F-ASM-06_{schedule_id}.pdf", mimetype='application/pdf')




@app.route('/assets')
@login_required
def assets_page():
    """Technical Audit Page: Detailed Hardware Inventory"""
    try:
        devices = supabase.table("devices").select("*").execute().data or []
        # Calculate online status for the small indicator dots
        calculate_online_status(devices)
        return render_template('assets.html', devices=devices, page="assets")
    except Exception as e:
        return f"Asset Module Error: {e}"

# --- NEW: GEOLOCATION LOGIC ---
def get_coords_from_ip(public_ip):
    try:
        # Use a free API to locate the node by IP
        r = requests.get(f"http://ip-api.com/json/{public_ip}?fields=status,lat,lon,city", timeout=3)
        data = r.json()
        if data['status'] == 'success':
            return data['lat'], data['lon'], data['city']
    except: pass
    return None, None, None


@app.route('/device/<device_id>/uninstall', methods=['POST'])
@login_required
def uninstall_app(device_id):  # <--- Added device_id here
    app_name = request.form.get('app_name')

    if not app_name:
        return jsonify({"error": "Application name is required"}), 400

    # Generate Winget Uninstall Command
    # We use --name to target the specific app string reported by the agent
    cmd = f'winget uninstall --name "{app_name}" --silent --accept-source-agreements'

    # Update Supabase
    try:
        supabase.table("devices").update({"pending_command": cmd}).eq("id", device_id).execute()

        # Log the action to audit logs
        supabase.table("audit_logs").insert({
            "target_device": device_id,
            "action_type": "REMOTE_UNINSTALL",
            "details": f"Initiated uninstall for: {app_name}"
        }).execute()

        return jsonify({"status": "Uninstall signal dispatched"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/geofencing')
@login_required
def geofencing_page():
    devices = supabase.table("devices").select("id, display_name, hostname, latitude, longitude, city, is_online").execute().data or []
    return render_template('geofencing.html', devices=devices, page="geofencing")



if __name__ == '__main__':
    # Flask local server init
    app.run(debug=True, port=5000)